"""
Blindseeker v1.0.0 - Export Module
====================================
Multi-format export system for forensics reports.
Supports: JSON, CSV, PDF, XML, HTML, XLSX
"""

import os
import json
import csv
import io
import logging
from datetime import datetime, timezone

logger = logging.getLogger('blindseeker.exporter')


class Exporter:
    """Multi-format scan result exporter."""
    
    def __init__(self, export_dir='exports'):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)
    
    def export(self, scan_data, fmt='json', filename=None, investigator='',
               case_id='', notes=''):
        """
        Export scan results in the specified format.
        
        Args:
            scan_data: ScanSession.to_dict() output
            fmt: Export format (json, csv, pdf, xml, html, xlsx)
            filename: Optional custom filename
            investigator: Investigator name for reports
            case_id: Case identifier for forensics tracking
            notes: Additional notes
        
        Returns:
            filepath or bytes depending on context
        """
        if not filename:
            ts = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
            username = scan_data.get('username', 'unknown')
            filename = f"blindseeker_{username}_{ts}"
        
        # Add metadata
        scan_data['export_metadata'] = {
            'exported_at': datetime.now(timezone.utc).isoformat(),
            'format': fmt,
            'tool': 'Blindseeker v1.0.0',
            'investigator': investigator,
            'case_id': case_id,
            'notes': notes
        }
        
        fmt = fmt.lower()
        
        if fmt == 'json':
            return self._export_json(scan_data, filename)
        elif fmt == 'csv':
            return self._export_csv(scan_data, filename)
        elif fmt == 'pdf':
            return self._export_pdf(scan_data, filename, investigator, case_id, notes)
        elif fmt == 'xml':
            return self._export_xml(scan_data, filename)
        elif fmt == 'html':
            return self._export_html(scan_data, filename)
        elif fmt == 'xlsx':
            return self._export_xlsx(scan_data, filename)
        else:
            raise ValueError(f"Unsupported format: {fmt}")
    
    def _export_json(self, data, filename):
        """Export as JSON."""
        filepath = os.path.join(self.export_dir, f"{filename}.json")
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        logger.info(f"Exported JSON: {filepath}")
        return filepath
    
    def _export_csv(self, data, filename):
        """Export as CSV."""
        filepath = os.path.join(self.export_dir, f"{filename}.csv")
        
        found = data.get('found', [])
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                'Platform', 'Username', 'URL', 'Status Code',
                'Response Time (ms)', 'Category', 'Found'
            ])
            
            # Found results
            for r in found:
                writer.writerow([
                    r.get('platform', ''),
                    r.get('username', ''),
                    r.get('url', ''),
                    r.get('status_code', ''),
                    r.get('response_time', ''),
                    r.get('category', ''),
                    'Yes'
                ])
            
            # Not found results
            for r in data.get('not_found', []):
                writer.writerow([
                    r.get('platform', ''),
                    r.get('username', ''),
                    r.get('url', ''),
                    r.get('status_code', ''),
                    r.get('response_time', ''),
                    r.get('category', ''),
                    'No'
                ])
        
        logger.info(f"Exported CSV: {filepath}")
        return filepath
    
    def _export_pdf(self, data, filename, investigator='', case_id='', notes=''):
        """Export as PDF forensics report."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph,
                Spacer, HRFlowable, PageBreak
            )
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            logger.error("reportlab not installed. Install with: pip install reportlab")
            # Fallback to JSON
            return self._export_json(data, filename)
        
        filepath = os.path.join(self.export_dir, f"{filename}.pdf")
        
        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            topMargin=0.5*inch, bottomMargin=0.5*inch,
            leftMargin=0.5*inch, rightMargin=0.5*inch
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Title'],
            fontSize=24, textColor=colors.HexColor('#00ff88'),
            spaceAfter=10
        )
        subtitle_style = ParagraphStyle(
            'Subtitle', parent=styles['Normal'],
            fontSize=10, textColor=colors.grey,
            spaceAfter=20
        )
        heading_style = ParagraphStyle(
            'CustomHeading', parent=styles['Heading2'],
            fontSize=14, textColor=colors.HexColor('#0088ff'),
            spaceBefore=15, spaceAfter=8
        )
        normal_style = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontSize=9, spaceAfter=4
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph("BLINDSEEKER v1.0.0", title_style))
        elements.append(Paragraph("Username Enumeration Forensics Report", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#00ff88')))
        elements.append(Spacer(1, 15))
        
        # Metadata
        elements.append(Paragraph("SCAN METADATA", heading_style))
        meta_data = [
            ['Field', 'Value'],
            ['Target Username', data.get('username', 'N/A')],
            ['Scan ID', data.get('scan_id', 'N/A')],
            ['Start Time', data.get('start_time', 'N/A')],
            ['End Time', data.get('end_time', 'N/A')],
            ['Duration', f"{data.get('duration_seconds', 0)} seconds"],
            ['Platforms Scanned', str(data.get('total_platforms', 0))],
            ['Profiles Found', str(data.get('found_count', 0))],
            ['Errors', str(data.get('error_count', 0))],
            ['Investigator', investigator or 'N/A'],
            ['Case ID', case_id or 'N/A'],
        ]
        
        meta_table = Table(meta_data, colWidths=[2*inch, 4.5*inch])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f5f5'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 20))
        
        # Found profiles
        found = data.get('found', [])
        if found:
            elements.append(Paragraph(f"PROFILES FOUND ({len(found)})", heading_style))
            
            table_data = [['#', 'Platform', 'Category', 'URL', 'Response Time']]
            for i, r in enumerate(found, 1):
                table_data.append([
                    str(i),
                    r.get('platform', 'Unknown'),
                    r.get('category', 'other').title(),
                    Paragraph(f"<link href='{r.get('url', '')}'>{r.get('url', 'N/A')[:50]}</link>", normal_style),
                    f"{r.get('response_time', 'N/A')} ms"
                ])
            
            found_table = Table(table_data, colWidths=[0.4*inch, 1.3*inch, 1*inch, 3.3*inch, 0.8*inch])
            found_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#006633')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#e8f5e9'), colors.white]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(found_table)
        
        elements.append(Spacer(1, 20))
        
        # Notes
        if notes:
            elements.append(Paragraph("INVESTIGATOR NOTES", heading_style))
            elements.append(Paragraph(notes, normal_style))
            elements.append(Spacer(1, 15))
        
        # Footer
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
        elements.append(Spacer(1, 5))
        footer_style = ParagraphStyle(
            'Footer', parent=styles['Normal'],
            fontSize=7, textColor=colors.grey, alignment=TA_CENTER
        )
        elements.append(Paragraph(
            f"Generated by Blindseeker v1.0.0 | {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')} | "
            "FOR AUTHORIZED LAW ENFORCEMENT AND SECURITY USE ONLY",
            footer_style
        ))
        
        doc.build(elements)
        logger.info(f"Exported PDF: {filepath}")
        return filepath
    
    def _export_xml(self, data, filename):
        """Export as XML."""
        filepath = os.path.join(self.export_dir, f"{filename}.xml")
        
        try:
            from dicttoxml import dicttoxml
            xml_bytes = dicttoxml(data, custom_root='blindseeker_report', attr_type=False)
        except ImportError:
            # Manual XML generation fallback
            xml_bytes = self._manual_xml(data).encode('utf-8')
        
        with open(filepath, 'wb') as f:
            f.write(xml_bytes)
        
        logger.info(f"Exported XML: {filepath}")
        return filepath
    
    def _manual_xml(self, data, root='blindseeker_report', indent=0):
        """Generate XML manually without dicttoxml."""
        lines = []
        if indent == 0:
            lines.append('<?xml version="1.0" encoding="UTF-8"?>')
        
        prefix = '  ' * indent
        
        if isinstance(data, dict):
            lines.append(f"{prefix}<{root}>")
            for key, value in data.items():
                safe_key = str(key).replace(' ', '_').replace('-', '_')
                lines.append(self._manual_xml(value, safe_key, indent + 1))
            lines.append(f"{prefix}</{root}>")
        elif isinstance(data, list):
            lines.append(f"{prefix}<{root}>")
            for item in data:
                lines.append(self._manual_xml(item, 'item', indent + 1))
            lines.append(f"{prefix}</{root}>")
        else:
            safe_val = str(data).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            lines.append(f"{prefix}<{root}>{safe_val}</{root}>")
        
        return '\n'.join(lines)
    
    def _export_html(self, data, filename):
        """Export as styled HTML report."""
        filepath = os.path.join(self.export_dir, f"{filename}.html")
        
        found = data.get('found', [])
        
        rows = ''
        for i, r in enumerate(found, 1):
            rows += f"""
            <tr>
                <td>{i}</td>
                <td>{r.get('platform', 'Unknown')}</td>
                <td><span class="badge">{r.get('category', 'other')}</span></td>
                <td><a href="{r.get('url', '#')}" target="_blank">{r.get('url', 'N/A')}</a></td>
                <td>{r.get('response_time', 'N/A')} ms</td>
            </tr>"""
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Blindseeker Report - {data.get('username', '')}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', sans-serif; background: #0a0a1a; color: #e0e0e0; padding: 40px; }}
        .container {{ max-width: 1000px; margin: 0 auto; }}
        h1 {{ color: #00ff88; font-size: 28px; margin-bottom: 5px; }}
        h2 {{ color: #0088ff; font-size: 18px; margin: 20px 0 10px; }}
        .subtitle {{ color: #888; margin-bottom: 20px; }}
        .meta {{ background: #111; border: 1px solid #222; border-radius: 8px; padding: 20px; margin: 20px 0; }}
        .meta-row {{ display: flex; justify-content: space-between; padding: 6px 0; border-bottom: 1px solid #1a1a2e; }}
        .meta-label {{ color: #888; }}
        .meta-value {{ color: #fff; font-weight: 600; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
        th {{ background: #006633; color: white; padding: 10px; text-align: left; font-size: 13px; }}
        td {{ padding: 8px 10px; border-bottom: 1px solid #1a1a2e; font-size: 12px; }}
        tr:hover {{ background: rgba(0,255,136,0.05); }}
        a {{ color: #00ff88; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        .badge {{ background: #1a1a2e; color: #0088ff; padding: 2px 8px; border-radius: 4px; font-size: 11px; }}
        .stats {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin: 20px 0; }}
        .stat-card {{ background: #111; border: 1px solid #222; border-radius: 8px; padding: 15px; text-align: center; }}
        .stat-value {{ font-size: 24px; font-weight: 700; color: #00ff88; }}
        .stat-label {{ font-size: 11px; color: #888; margin-top: 4px; }}
        .footer {{ text-align: center; color: #444; font-size: 11px; margin-top: 40px; padding-top: 20px; border-top: 1px solid #222; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>BLINDSEEKER v1.0.0</h1>
        <p class="subtitle">Username Enumeration Forensics Report</p>
        
        <div class="stats">
            <div class="stat-card"><div class="stat-value">{data.get('total_platforms', 0)}</div><div class="stat-label">Platforms Scanned</div></div>
            <div class="stat-card"><div class="stat-value">{data.get('found_count', 0)}</div><div class="stat-label">Profiles Found</div></div>
            <div class="stat-card"><div class="stat-value">{data.get('duration_seconds', 0)}s</div><div class="stat-label">Scan Duration</div></div>
            <div class="stat-card"><div class="stat-value">{data.get('error_count', 0)}</div><div class="stat-label">Errors</div></div>
        </div>
        
        <div class="meta">
            <h2>Scan Details</h2>
            <div class="meta-row"><span class="meta-label">Username</span><span class="meta-value">{data.get('username', 'N/A')}</span></div>
            <div class="meta-row"><span class="meta-label">Scan ID</span><span class="meta-value">{data.get('scan_id', 'N/A')}</span></div>
            <div class="meta-row"><span class="meta-label">Start Time</span><span class="meta-value">{data.get('start_time', 'N/A')}</span></div>
            <div class="meta-row"><span class="meta-label">End Time</span><span class="meta-value">{data.get('end_time', 'N/A')}</span></div>
        </div>
        
        <h2>Profiles Found ({len(found)})</h2>
        <table>
            <thead><tr><th>#</th><th>Platform</th><th>Category</th><th>Profile URL</th><th>Response</th></tr></thead>
            <tbody>{rows}</tbody>
        </table>
        
        <div class="footer">
            Generated by Blindseeker v1.0.0 | FOR AUTHORIZED LAW ENFORCEMENT AND SECURITY USE ONLY
        </div>
    </div>
</body>
</html>"""
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html)
        
        logger.info(f"Exported HTML: {filepath}")
        return filepath
    
    def _export_xlsx(self, data, filename):
        """Export as Excel XLSX."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return self._export_csv(data, filename)
        
        filepath = os.path.join(self.export_dir, f"{filename}.xlsx")
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="006633", end_color="006633", fill_type="solid")
        title_font = Font(bold=True, size=14, color="00FF88")
        
        ws_summary['A1'] = 'BLINDSEEKER v1.0.0 - Scan Report'
        ws_summary['A1'].font = title_font
        
        meta_fields = [
            ('Username', data.get('username', '')),
            ('Scan ID', data.get('scan_id', '')),
            ('Start Time', data.get('start_time', '')),
            ('End Time', data.get('end_time', '')),
            ('Duration (s)', data.get('duration_seconds', 0)),
            ('Platforms Scanned', data.get('total_platforms', 0)),
            ('Profiles Found', data.get('found_count', 0)),
            ('Errors', data.get('error_count', 0)),
        ]
        
        for i, (label, value) in enumerate(meta_fields, 3):
            ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=i, column=2, value=str(value))
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 50
        
        # Found profiles sheet
        ws_found = wb.create_sheet("Found Profiles")
        headers = ['#', 'Platform', 'Category', 'URL', 'Status Code', 'Response Time (ms)']
        
        for col, header in enumerate(headers, 1):
            cell = ws_found.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for i, r in enumerate(data.get('found', []), 1):
            ws_found.cell(row=i+1, column=1, value=i)
            ws_found.cell(row=i+1, column=2, value=r.get('platform', ''))
            ws_found.cell(row=i+1, column=3, value=r.get('category', ''))
            ws_found.cell(row=i+1, column=4, value=r.get('url', ''))
            ws_found.cell(row=i+1, column=5, value=r.get('status_code', ''))
            ws_found.cell(row=i+1, column=6, value=r.get('response_time', ''))
        
        ws_found.column_dimensions['B'].width = 20
        ws_found.column_dimensions['C'].width = 15
        ws_found.column_dimensions['D'].width = 60
        
        # All Results sheet
        ws_all = wb.create_sheet("All Results")
        all_headers = ['Platform', 'Found', 'Category', 'URL', 'Status Code', 'Response Time', 'Error']
        
        for col, header in enumerate(all_headers, 1):
            cell = ws_all.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        all_results = data.get('found', []) + data.get('not_found', []) + data.get('errors', [])
        for i, r in enumerate(all_results, 1):
            ws_all.cell(row=i+1, column=1, value=r.get('platform', ''))
            ws_all.cell(row=i+1, column=2, value='Yes' if r.get('found') else 'No')
            ws_all.cell(row=i+1, column=3, value=r.get('category', ''))
            ws_all.cell(row=i+1, column=4, value=r.get('url', ''))
            ws_all.cell(row=i+1, column=5, value=r.get('status_code', ''))
            ws_all.cell(row=i+1, column=6, value=r.get('response_time', ''))
            ws_all.cell(row=i+1, column=7, value=r.get('error', ''))
        
        wb.save(filepath)
        logger.info(f"Exported XLSX: {filepath}")
        return filepath
    
    def export_bytes(self, scan_data, fmt='json', investigator='', case_id='', notes=''):
        """Export to bytes (for HTTP response download)."""
        ts = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        username = scan_data.get('username', 'unknown')
        filename = f"blindseeker_{username}_{ts}"
        
        filepath = self.export(scan_data, fmt, filename, investigator, case_id, notes)
        
        with open(filepath, 'rb') as f:
            content = f.read()
        
        return content, os.path.basename(filepath)
